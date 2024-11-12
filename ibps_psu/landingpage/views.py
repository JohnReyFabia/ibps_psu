
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.template import loader
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_page
from django.core.cache import cache
import time
from django.db.models.signals import post_save, pre_save, post_delete
from django.dispatch import receiver

# User authentication
from django.contrib.auth.models import User

# Excel/CVC reader
import pandas as pd

# Download to excel
import datetime
import openpyxl
import os
import io
import pytz
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

# regular expressions
import re

# For queries
from django.db.models import Q, Count

# Survey Questions Display
import json

# Keyword Extraction
from rake_nltk import Rake
import yake
import numpy as np

# Messages
from django.db import IntegrityError
from django.contrib import messages

# Login authentication
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect
from django.core.exceptions import ValidationError

# ChangePassword authentication
from django.contrib.auth import update_session_auth_hash

# ForgotPassword
from django.conf import settings
from django.core.mail import send_mail
import uuid
import hashlib
from datetime import timedelta
from django.utils import timezone

#  Burnout assessment forms and models import
from django.db import models
from burnout_assessment.models import (
    BurnoutProfile,
    Student,
    Counselor,
    College,
    Program,
    Assessment,
    StudentSurveyQuestion,
    SurveyQuestion,
    ForgotPasswordRequest,
    SurveyQuestionChoice,
)

from .forms import (
    StudentForm,
    CounselorForm,
    LoginForm,
    UpdateProfileForm,
    ChangePasswordForm,
    ForgotPassForm,
    ChangePassForm,
)

# for display of charts
from .utils import (
    check_for_new_assessment,
    get_keywords_and_counts_for_students,
    calculate_age_range_distribution,
    get_overall_burnout_counts,
    get_profile_gender_counts,
    get_program_info,
)


def test(request):
    pass
    # def insert_row(d):

    #     possible_answers = {
    #         0: SurveyQuestionChoice.objects.get(value=0),
    #         1: SurveyQuestionChoice.objects.get(value=1),
    #         2: SurveyQuestionChoice.objects.get(value=2),
    #         3: SurveyQuestionChoice.objects.get(value=3),
    #         4: SurveyQuestionChoice.objects.get(value=4),
    #         5: SurveyQuestionChoice.objects.get(value=5),
    #         6: SurveyQuestionChoice.objects.get(value=6),
    #     }
    #     if User.objects.filter(email=d["Email Address"]).exists():
    #         user = User.objects.get(email=d["Email Address"])
    #     else:
    #         form = StudentForm(
    #             {
    #                 "email": d["Email Address"],
    #                 "student_id": d["Student ID (0000-00-0000)"],
    #                 "program": d["Program"],
    #                 "password": "1234",
    #                 "confirm_password": "1234",
    #             }
    #         )
    #         if form.is_valid():
    #             user = form.save()
    #         else:
    #             raise Exception(f"Error creating user: {form.errors}")

    #     student = Student.objects.get(account=user)

    #     survey_form = StudentSurveyForm(
    #         {
    #             "question_15": possible_answers.get(d["CY4_SCORE"]),
    #             "question_14": possible_answers.get(d["CY3_SCORE"]),
    #             "question_13": possible_answers.get(d["CY2_SCORE"]),
    #             "question_12": possible_answers.get(d["CY1_SCORE"]),
    #             "question_11": possible_answers.get(d["EX5_SCORE"]),
    #             "question_10": possible_answers.get(d["EX4_SCORE"]),
    #             "question_9": possible_answers.get(d["EX3_SCORE"]),
    #             "question_8": possible_answers.get(d["EX2_SCORE"]),
    #             "question_7": possible_answers.get(d["EX1_SCORE"]),
    #             "question_6": possible_answers.get(d["EF5_SCORE"]),
    #             "question_5": possible_answers.get(d["EF4_SCORE"]),
    #             "question_4": possible_answers.get(d["EF3_SCORE"]),
    #             "question_3": possible_answers.get(d["EF2_SCORE"]),
    #             "question_2": possible_answers.get(d["EF1_SCORE"]),
    #             "question_1": possible_answers.get(d["EF6_SCORE"]),
    #         }
    #     )

    #     if survey_form.is_valid():
    #         survey = survey_form.save_to_assessment(student)
    #     else:
    #         raise Exception(f"Error saving survey: {survey_form.errors}")

    #     print(f"Survery successfully saved survey: {survey_form.errors}")

    # if request.method == "POST":
    #     data = request.POST.get('data')

    #     print(f'data recieved at 0: {data[0]}')

    # return JsonResponse({'data': data})


# ----------------------------------------------------------------
#  LANDING PAGE
# ----------------------------------------------------------------


def home(request):
    profile = BurnoutProfile.objects.exclude(id=9).order_by("-id")
    return render(request, "index.html", {"profile": profile})


# ----------------------------------------------------------------
#  AUTHENTICATION FUNCTIONS
# ----------------------------------------------------------------

#Landing page send email 
def send_email_form(request):
    success = False
    text = None

    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')

        # Print the received data
        print(f"Name: {name}")
        print(f"Email: {email}")
        print(f"Subject: {subject}")
        print(f"Message: {message}")

        # Send an email with the inputted email as the sender
        send_mail(
            f'Subject: {subject}',
            f'Message from {name} ({email}):\n\n{message}',
            email, 
            ['ibpspsu2023@gmail.com'],  # Recipient email address
            fail_silently=False,
        )

        text = "Your message has been sent. Thank you!"
        success = True
    else:
        text = "There was an error sending your message. Please try again later."
        success = False
    
    return render(request, "index.html", {"success": success, "text": text})


# Student Account Registration
def student_register(request):
    success = False
    text = None
    registerform = StudentForm()

    if request.method == "POST":
        registerform = StudentForm(request.POST)
        if registerform.is_valid():
            # email = registerform.cleaned_data.get("email")
            student_id = registerform.cleaned_data.get("student_id")

            # Check if a student with the same student_id already exists
            if Student.objects.filter(student_id=student_id).exists():
                messages.error(request, "A user with the same Student ID already exists.")
                success = False
            else:
                try:
                    registerform.save()
                    text = "User Created Successfully!"
                    success = True
                    # return redirect("landingpage:login")

                    # # Save the form to get access to the instance
                    # student = registerform.save(commit=False)

                    # # Check if the email is valid
                    # if registerform.is_valid_email(email):
                    #     text = "Verification email sent. Please check your inbox and enter the code."
                    #     success = True
                    # else:
                    #     messages.error(request, "Invalid email address.")
                    #     success = False
                except IntegrityError as e:
                    messages.error(request, "Form is not valid")
                    print(f"IntegrityError: {e}")
                    success = False
        else:
            messages.error(request, "Form is not valid")
            success = False


    context = {
        "success": success,
        "text": text,
        "mode": "sign-up-mode",
        "form": {
            "register": registerform,
            "login": LoginForm(),
            "forgot_password": ForgotPassForm(),
        },
    }

    return render(request, "account/sign-in-up.html", context)


# Counselor Account Registration
def counselor_register(request):
    success = False
    text = None
    register_form = CounselorForm()

    if request.method == "POST":
        register_form = CounselorForm(request.POST)
        if register_form.is_valid():

            try:
                register_form.save()
                text = "User Created Successfully!"
                success = True

            except IntegrityError as e:
                messages.error(request, "Form is not valid")
                print(f"IntegrityError: {e}")
                success = False
        else:
            messages.error(request, "Form is not valid")
            success = False


    context = {
        "success": success,
        "text": text,
        "mode": "sign-up-mode",
        "form": {
            "register_form": register_form,
            "login": LoginForm(),
            "forgot_password": ForgotPassForm(),
        },
    }

    return render(request, "account/sign-in-up.html", context)


# User login
def user_login(request):
    loginform = LoginForm(request.POST)

    if request.method == "POST":
        if loginform.is_valid():
            user_log = loginform.cleaned_data["user_log"]
            print(user_log)
            user_pass = loginform.cleaned_data["user_pass"]

            user = authenticate(request, username=user_log, password=user_pass)

            if user is not None:
                login(request, user)
                if request.user.is_superuser:
                    # Redirect to superuser admin dashboard
                    return redirect("landingpage:admin-dashboard")
                elif not request.user.is_superuser and not any(char.isdigit() for char in user.email):
                    # Redirect to counserlor dashboard
                    return redirect("landingpage:counselor-dashboard")
                else:
                    # # Set the session variable to indicate a new login
                    # request.session["just_logged_in"] = True
                    # Redirect to student dashboard
                    return redirect("landingpage:student-dashboard")
            else:
                loginform.add_error(None, "Invalid username or password")
                # messages.error(request, 'Authentication failed')
                print(user_log)

    else:
        loginform = LoginForm()

    context = {
        "mode": "",
        "form": {
            "register": StudentForm(),
            "login": loginform,
            "forgot_password": ForgotPassForm(),
        },
    }

    return render(request, "account/sign-in-up.html", context)


# User Logout
def user_logout(request):
    cache_key = f"user_dashboard:{request.user.id}"
    cache.delete(cache_key)
    logout(request)
    return redirect("landingpage:home")


def forget_password(request):
    forgotpassform = ForgotPassForm(request.POST or None)
    success = False
    text = None

    try:
        if request.method == "POST":
            email = request.POST.get("email")

            if not User.objects.filter(email=email).exists():
                forgotpassform.add_error("email", "No user found with this email address.")
                messages.error(request, "No user found with this email address.")
                return render(
                    request,
                    "forget-pass.html",
                    {
                        "form": {
                            "register": StudentForm(),
                            "login": LoginForm(),
                            "forgot_password": forgotpassform,
                        }
                    },
                )

            user_obj = User.objects.get(email=email)
            token = str(uuid.uuid4())
            hashed_token = hashlib.sha256(token.encode()).hexdigest()
            expiration_time = timezone.now() + timedelta(minutes=10)

            # Checking for ForgotPasswordRequest for the user if exists
            try:
                profile_obj = ForgotPasswordRequest.objects.get(user=user_obj)
                profile_obj.forget_pass_token = hashed_token
                profile_obj.expiration_time = expiration_time
            except ForgotPasswordRequest.DoesNotExist:
                # Create a new object if not exists
                profile_obj = ForgotPasswordRequest(
                    user=user_obj,
                    forget_pass_token=hashed_token,
                    expiration_time=expiration_time,
                )

            profile_obj.save()
            send_forgetpass_email(user_obj.email, hashed_token, request)
            text = "Successfully sent an email!"
            messages.success(request, "Successfully sent an email!")
            success = True

    except Exception as e:
        print(e)

    context = {
        "success": success,
        "text": text,
        "mode": "forgot-pass-mode",
        "form": {
            "register": StudentForm(),
            "login": LoginForm(),
            "forgot_password": forgotpassform,
        },
    }

    return render(request, "account/sign-in-up.html", context)


def reset_password(request, hashed_token):
    form = ChangePassForm(request.POST or None)
    context = {}
    success = False
    try:
        profile_obj = ForgotPasswordRequest.objects.filter(
            forget_pass_token=hashed_token
        ).first()

        # Check if the token is valid and hasn't expired
        if not profile_obj or profile_obj.is_expired():
            # msg = 'The reset password link is invalid or has expired. Please try again.'
            html_template = loader.get_template("page-404.html")
            return HttpResponse(html_template.render(context, request))
            # return redirect('/forget_pass/')

        context = {"user_id": profile_obj.user.id}

        if request.method == "POST":
            new_password1 = request.POST.get("new_password1")
            new_password2 = request.POST.get("new_password2")
            user_id = request.POST.get("user_id")

            if user_id is None:
                messages.error(request, "No user found.")
                return redirect(f"/reset_password/")

            if new_password1 != new_password2:
                form.add_error(None, "Password does not match!")
                messages.error(request, "Password does not match!")

                # return redirect(f'/change_password/{hashed_token}/')
            else:
                if not request.user.is_superuser and not any(
                    char.isdigit() for char in profile_obj.user.email
                ):
                    counselor = Counselor.objects.get(account__pk=context["user_id"])
                    counselor.account.set_password(new_password2)
                    counselor.account.save()
                    profile_obj.delete()
                    success = True
                else:
                    student = Student.objects.get(account__pk=context["user_id"])
                    student.account.set_password(new_password2)
                    student.account.save()
                    profile_obj.delete()
                    success = True
                # return redirect('/login/')

    except Exception as e:
        print(e)
    return render(
        request,
        "account/reset-pass.html",
        {"form": form, "context": context, "success": success},
    )


def send_forgetpass_email(email, hashed_token, request):
    subject = "IBPS PSU Password Reset"
    reset_link = f'{request.META["HTTP_ORIGIN"]}/reset_password/{hashed_token}/'
    message = f'''
    <html>
    <body>
        <p>Hello,</p>
        <p>We received a request to reset your password for IBPS PSU.</p>
        <p>To reset your password, please click on the following button:</p>
        <p><a href="{reset_link}" style=" background-color: orange; color: white; padding: 10px 20px; text-align: center; text-decoration: none; display: inline-block; font-size: 12px; margin: 4px 2px; cursor: pointer; border-radius: 5px;">Reset Password</a></p>
        <p>If you did not initiate this request, you can safely ignore this email. Your password will not be changed.</p>
    </body>
    </html>
    '''
   
    email_from = settings.EMAIL_HOST_USER
    recipient = [email]
    send_mail(subject, '', email_from, recipient, html_message=message)
    return True


# User Change Password
@login_required(login_url="/login/")
def change_password(request):
    student = Student.objects.filter(email=request.user.email).first()
    if request.method == "POST":
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            
            if not student.is_password_updated:
                student.is_password_updated = True
                student.save()

            update_session_auth_hash(request, user)  # Important to update the session hash
            messages.success(request, "Your password has been changed successfully.")
            if request.user.is_superuser:
                return redirect("landingpage:admin-dashboard")
            elif not request.user.is_superuser and not any(char.isdigit() for char in user.email):
                # Redirect to counserlor dashboard
                return redirect("landingpage:counselor-dashboard")
            else:
                return redirect("landingpage:edit-profile")
        # else:
        #     messages.error(request, 'There was an error changing your password.')
    else:
        form = ChangePasswordForm(request.user)

    student_id = student.student_id if student else None

    return render(
        request,
        "account/change-password.html",
        {"form": form, "student_id": student_id},
    )


# ----------------------------------------------------------------
#  ADMIN SIDE FUNCTIONS
# ----------------------------------------------------------------


@receiver(post_save, sender=Student)
@receiver(post_delete, sender=Student)
def update_cache_on_change(sender, instance, **kwargs):
    # Invalidate or update the relevant cache keys
    cache.delete('last_assessment_timestamp')
    cache.delete('admin_dashboard_data')
    cache.delete('counselor_dashboard_data')


ADMIN_CACHE_PREFIX = 'admin_dashboard_data'

# Admin Dashboard
@login_required(login_url="/login/")
def admin_dashboard(request):
    start_time = time.time()

    # Extract selected college from URL parameters
    selected_college = request.GET.get("college_name", None)

    cache_key = f'{ADMIN_CACHE_PREFIX}_{selected_college}'

    # Check if the cache exists
    cached_data = cache.get(cache_key)

    new_assessment = check_for_new_assessment()

    if not cached_data or new_assessment:

        if selected_college:
            college_programs_start_time = time.time()
            # Get all programs associated with the selected college
            college_programs = Program.objects.filter(college__college_name=selected_college)
            college_programs_end_time = time.time()
            print(f"College Programs Query Time: {college_programs_end_time - college_programs_start_time} seconds")

            students_start_time = time.time()
            # Get students associated with the selected college
            students = Student.objects.filter(program__college__college_name=selected_college)
            students_end_time = time.time()
            print(f"Students Query Time: {students_end_time - students_start_time} seconds")

            total_students_start_time = time.time()
            total_students = Student.objects.filter(program__college__college_name=selected_college).count()
            total_students_end_time = time.time()
            print(f"Total Students Query Time: {total_students_end_time - total_students_start_time} seconds")

            total_student_takers_start_time = time.time()
            total_student_takers = Assessment.objects.select_related("student").filter(student__program__college__college_name=selected_college).count()
            total_student_takers_end_time = time.time()
            print(f"Total Student Takers Query Time: {total_student_takers_end_time - total_student_takers_start_time} seconds")

            student_assessments_start_time = time.time()
            student_assessments = Assessment.objects.select_related("student").filter(student__program__college__college_name=selected_college).order_by('-pk')[:8]
            student_assessments_end_time = time.time()
            print(f"Student Assessments Query Time: {student_assessments_end_time - student_assessments_start_time} seconds")

            burnout_profiles_start_time = time.time()
            burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
            burnout_profiles_end_time = time.time()
            print(f"Burnout Profiles Query Time: {burnout_profiles_end_time - burnout_profiles_start_time} seconds")

            keywords_start_time = time.time()
            top_keywords = get_keywords_and_counts_for_students(students)
            keywords_end_time = time.time()
            print(f"Keywords Query Time: {keywords_end_time - keywords_start_time} seconds")

            age_data_start_time = time.time()
            age_data_per_profile = calculate_age_range_distribution(students, burnout_profiles)
            age_data_end_time = time.time()
            print(f"Age Data Query Time: {age_data_end_time - age_data_start_time} seconds")

            program_info_start_time = time.time()
            program_info = get_program_info(college_programs)
            program_info_end_time = time.time()
            print(f"Program Info Query Time: {program_info_end_time - program_info_start_time} seconds")

            overall_burnout_counts_start_time = time.time()
            overall_burnout_counts = get_overall_burnout_counts(selected_college)
            overall_burnout_counts_end_time = time.time()
            print(f"Overall Burnout Counts Query Time: {overall_burnout_counts_end_time - overall_burnout_counts_start_time} seconds")

            gender_counts_start_time = time.time()
            profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)
            gender_counts_end_time = time.time()
            print(f"Gender Counts Query Time: {gender_counts_end_time - gender_counts_start_time} seconds")

            chart_labels_start_time = time.time()
            chart_labels = list(profile_gender_counts.keys())
            chart_labels_end_time = time.time()
            print(f"Chart Labels Query Time: {chart_labels_end_time - chart_labels_start_time} seconds")

            male_data_start_time = time.time()
            male_data = [profile_gender_counts[profile]["Male"] for profile in chart_labels]
            male_data_end_time = time.time()
            print(f"Male Data Query Time: {male_data_end_time - male_data_start_time} seconds")

            female_data_start_time = time.time()
            female_data = [profile_gender_counts[profile]["Female"] for profile in chart_labels]
            female_data_end_time = time.time()
            print(f"Female Data Query Time: {female_data_end_time - female_data_start_time} seconds")

        else:
            all_colleges_start_time = time.time()
            # If no specific college is selected, fetch information from all colleges
            college_programs_start_time = time.time()
            college_programs = Program.objects.all()
            college_programs_end_time = time.time()
            print(f"All Colleges - College Programs Query Time: {college_programs_end_time - college_programs_start_time} seconds")

            students_start_time = time.time()
            students = Student.objects.all()
            students_end_time = time.time()
            print(f"All Colleges - Students Query Time: {students_end_time - students_start_time} seconds")

            total_students_start_time = time.time()
            total_students = Student.objects.all().count()
            total_students_end_time = time.time()
            print(f"All Colleges - Total Students Query Time: {total_students_end_time - total_students_start_time} seconds")

            total_student_takers_start_time = time.time()
            total_student_takers = Assessment.objects.select_related("student").count()
            total_student_takers_end_time = time.time()
            print(f"All Colleges - Total Student Takers Query Time: {total_student_takers_end_time - total_student_takers_start_time} seconds")

            student_assessments_start_time = time.time()
            student_assessments = Assessment.objects.select_related("student").order_by('-pk')[:8]
            student_assessments_end_time = time.time()
            print(f"All Colleges - Student Assessments Query Time: {student_assessments_end_time - student_assessments_start_time} seconds")

            burnout_profiles_start_time = time.time()
            burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
            burnout_profiles_end_time = time.time()
            print(f"All Colleges - Burnout Profiles Query Time: {burnout_profiles_end_time - burnout_profiles_start_time} seconds")

            keywords_start_time = time.time()
            top_keywords = get_keywords_and_counts_for_students(students)
            keywords_end_time = time.time()
            print(f"All Colleges - Keywords Query Time: {keywords_end_time - keywords_start_time} seconds")

            age_data_start_time = time.time()
            age_data_per_profile = calculate_age_range_distribution(students, burnout_profiles)
            age_data_end_time = time.time()
            print(f"All Colleges - Age Data Query Time: {age_data_end_time - age_data_start_time} seconds")

            program_info_start_time = time.time()
            program_info = get_program_info(college_programs)
            program_info_end_time = time.time()
            print(f"All Colleges - Program Info Query Time: {program_info_end_time - program_info_start_time} seconds")

            overall_burnout_counts_start_time = time.time()
            overall_burnout_counts = get_overall_burnout_counts()
            overall_burnout_counts_end_time = time.time()
            print(f"All Colleges - Overall Burnout Counts Query Time: {overall_burnout_counts_end_time - overall_burnout_counts_start_time} seconds")

            gender_counts_start_time = time.time()
            profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)
            gender_counts_end_time = time.time()
            print(f"All Colleges - Gender Counts Query Time: {gender_counts_end_time - gender_counts_start_time} seconds")

            chart_labels_start_time = time.time()
            chart_labels = list(profile_gender_counts.keys())
            chart_labels_end_time = time.time()
            print(f"All Colleges - Chart Labels Query Time: {chart_labels_end_time - chart_labels_start_time} seconds")

            male_data_start_time = time.time()
            male_data = [profile_gender_counts[profile]["Male"] for profile in chart_labels]
            male_data_end_time = time.time()
            print(f"All Colleges - Male Data Query Time: {male_data_end_time - male_data_start_time} seconds")

            female_data_start_time = time.time()
            female_data = [profile_gender_counts[profile]["Female"] for profile in chart_labels]
            female_data_end_time = time.time()
            print(f"All Colleges - Female Data Query Time: {female_data_end_time - female_data_start_time} seconds")

            all_colleges_end_time = time.time()
            print(f"All Colleges - Total Execution Time: {all_colleges_end_time - all_colleges_start_time} seconds")


        context = {
            "program_info": program_info,
            "selected_college": selected_college,
            "total_student_takers": total_student_takers,
            "total_students": total_students,
            "student_assessments": student_assessments,
            "overall_burnout_counts": overall_burnout_counts,
            "chart_labels": chart_labels,
            "male_data": male_data,
            "female_data": female_data,
            "top_keywords": top_keywords,
            "age_data_per_profile": age_data_per_profile,
        }

        cache.set(cache_key, context)

    else:
        context = cached_data
    
    end_time = time.time()
    print(f"Total Execution Time: {end_time - start_time} seconds")

    # return render(request, "admin/admin-dashboard.html", context)

    response = render(request, "admin/admin-dashboard.html", context)

    response['Cache-Control'] = 'max-age=' + str(6 * 30 * 24 * 60 * 60)  # Cache for 6 months
    response['Expires'] = (timezone.now() + timedelta(days=180)).strftime('%a, %d %b %Y %H:%M:%S GMT')

    return response


# Admin college list
@login_required(login_url="/login/")
def admin_college_list(request, code):
    college = get_object_or_404(College, code=code)
    students = Student.objects.filter(program__college=college)
    current_code = code 
    print(current_code)

    search = request.GET.get("q", "")

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(email__icontains=search)
            | Q(program__program_name__icontains=search)
            | Q(assessment__profile__profile__icontains=search)
        ).distinct()

    student_assessments = (
        Assessment.objects.select_related("student")
        .filter(student__in=students)
        .order_by("-id")
    )

    context = {
        "college": college,
        "student_assessments": student_assessments,
        "current_code": current_code,
        "search": search,
    }

    return render(request, "admin/admin-college.html", context)


# Admin student takers list
@login_required(login_url="/login/")
def admin_student_takers_list(request):
    selected_college = request.GET.get("college_name", None)
    search = request.GET.get("q", "")
    students = Student.objects.all()

    if selected_college:
        students = students.filter(program__college__college_name=selected_college)

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(email__icontains=search)
            | Q(program__program_name__icontains=search)
            | Q(assessment__profile__profile__icontains=search)
        )

    student_assessments = (
        Assessment.objects.select_related("student")
        .filter(student__in=students)
        .order_by("-id")
    )

    context = {
        "student_assessments": student_assessments,
        "search": search,
    }

    return render(request, "admin/admin-student-takers.html", context)



# Admin student list
@login_required(login_url="/login/")
def admin_student_list(request):
    selected_college = request.GET.get("college_name", None)
    print(selected_college)

    search = request.GET.get("q", "")

    if selected_college:
        students = Student.objects.filter(program__college__college_name=selected_college, assessment_exists=False).order_by("-id")
    else:
        students = Student.objects.filter(assessment_exists=False).order_by("-id")

    if search:
        students = students.filter(
            Q(student_id__icontains=search) |
            Q(email__icontains=search) |
            Q(program__program_name__icontains=search)
        ).distinct()

    context = {
        "students": students,
        "search": search,
    }

    return render(request, "admin/admin-student-list.html", context)

# Admin all student list
@login_required(login_url="/login/")
def admin_all_student_list(request):
    search = request.GET.get("q", "")

    students = Student.objects.all().order_by("-id")

    if search:
        students = students.filter(
            Q(student_id__icontains=search) |
            Q(email__icontains=search) |
            Q(program__program_name__icontains=search)
        ).distinct()

    context = {
        "students": students,
        "search": search,
    }

    return render(request, "admin/admin-all-student-list.html", context)

# Admin upload students
def upload_students_file(file):
    df = pd.read_excel(file)  # For Excel files
    # df = pd.read_csv(file)  # For CSV files
    
    # Get the program ID by mapping
    program_id_mapping = get_program_id_mapping()
    
    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        program_value = row['PROGRAM'] 
        
        # Get the corresponding program ID from the mapping
        program_id = program_id_mapping.get(program_value)
        
        if program_id is not None:
            # Extract other student data from the Excel file
            email = row['STUDENT EMAIL']
            student_id = row['STUDENT ID']
            contact_number = row['CONTACT NUMBER']
            last_name = row['LAST NAME']
            first_name = row['FIRST NAME']
            middle_name = row['MIDDLE NAME']
            gender = row['GENDER']
            birthdate = row['BIRTHDATE']
            civil_status = row['CIVIL STATUS']
            
            user = User.objects.create_user(username=student_id.replace('-', ''), email=email, password=last_name.lower(), last_name=last_name, first_name=first_name)
            
            student = Student.objects.create(
                account=user,
                student_id=student_id,
                email=email,
                contact_number=contact_number,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                gender=gender,
                birthdate=birthdate,
                civil_status=civil_status,
                program_id=program_id 
            )
            student.save()

            calculate_age(student.student_id)
            
            
        else:
            print(f"No matching program found for value: {program_value}")


def upload_students(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        upload_students_file(file)
        return JsonResponse({'message': 'Students added successfully'}, status=200)
    return JsonResponse({'error': 'Invalid request'}, status=400)


# def update_assessment_exists(request):
#     try:
#         # Update students with IDs ranging from 764 to 1000
#         Student.objects.filter(id__range=(764, 1000)).update(assessment_exists=True)
#         return HttpResponse("Assessment exists updated successfully for students with IDs 764 to 1000.")
#     except Exception as e:
#         return HttpResponse(f"An error occurred: {str(e)}")

# ----------------------------------------------------------------
#  ADMIN DOWNLOAD TO EXCEL FUNCTION
# ----------------------------------------------------------------

def admin_download_excel(request):
    file_path = os.path.join(settings.STATIC_ROOT, 'report_temp', 'excel_dashboard.xlsx')

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    selected_college = request.GET.get("college_name", None)
    print(selected_college)

    # Define the profiles variable with the profile names
    profiles = ['Engaged', 'Disengaged_and_Ineffective', 'Overextended_and_Ineffective', 'Overextended_and_Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned_out']
    b_profiles = ['Engaged', 'Disengaged and Ineffective', 'Overextended and Ineffective', 'Overextended and Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned Out']
    profile_per_prog = ['Engaged', 'Disengaged_and_Ineffective', 'Overextended_and_Ineffective', 'Overextended_and_Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned_Out']

    # Age distribution data for selected profiles
    selected_profiles = ['Engaged', 'Disengaged and Ineffective', 'Overextended and Ineffective', 'Overextended and Disengaged']

    # Starting columns for each profile
    start_columns = {'Engaged': 'K', 'Disengaged and Ineffective': 'O', 'Overextended and Ineffective': 'S', 'Overextended and Disengaged': 'W'}
    start_row_age = 17

    # Age distribution data for remaining profiles (Ineffective, Disengaged, Overextended, Burned Out)
    remaining_profiles = ['Ineffective', 'Disengaged', 'Overextended', 'Burned Out']
    start_rows_remaining = 20

    # Starting columns for remaining profiles
    start_columns_remaining = {'Ineffective': 'K', 'Disengaged': 'O', 'Overextended': 'S', 'Burned Out': 'W'}

    if selected_college:
        college_programs = Program.objects.filter(college__college_name=selected_college)
        students = Student.objects.filter(program__college__college_name=selected_college)
        total_students = Student.objects.filter(program__college__college_name=selected_college).count()
        total_student_takers = Assessment.objects.select_related("student").filter(
            student__program__college__college_name=selected_college).count()
        student_assessments = Assessment.objects.select_related("student").filter(
            student__program__college__college_name=selected_college).order_by('-created_date')
        burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
        age_data_per_profile = calculate_age_range_distribution(students, burnout_profiles)
        program_info = get_program_info(college_programs)
        overall_burnout_counts = get_overall_burnout_counts(selected_college)
        profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)

        #-----------------------------------------------------------------
        # Populate to excel
        # ----------------------------------------------------------------
        
        # College title
        ws['B8'] = selected_college  
        # Total student takers count
        ws['F8'] = total_student_takers
        # Total number of students count
        ws['G8'] = total_students
        # Overall Burnout Profiles count 
        get_overall_profile_counts(ws, overall_burnout_counts)
        # Burnout Profiles count per program
        get_program_with_burnout_count(ws, program_info, profile_per_prog)
        # Burnout Profiles count per gender
        get_profile_counts_per_gender(ws, profiles, profile_gender_counts)
        # Burnout Profiles count per age
        get_profile_count_per_age(ws, selected_profiles, start_columns, start_row_age, remaining_profiles, start_rows_remaining, start_columns_remaining, age_data_per_profile)
        # Create new sheet for student assessment list
        create_student_assessments_sheet(wb, student_assessments)
        # Create separate sheets for list of students according to profile
        create_profile_sheets(wb, b_profiles, selected_college)

    else:
        college_programs = Program.objects.all()
        students = Student.objects.all()
        total_students = Student.objects.all().count()
        total_student_takers = Assessment.objects.select_related("student").count()
        student_assessments = Assessment.objects.select_related("student").order_by('-created_date')
        burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
        age_data_per_profile = calculate_age_range_distribution(students, burnout_profiles)
        program_info = get_program_info(college_programs)
        overall_burnout_counts = get_overall_burnout_counts()
        profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)

        #-----------------------------------------------------------------
        # Populate to excel
        # ----------------------------------------------------------------

        # College/School Title
        ws['B8'] = 'Palawan State University'
        # Total student takers count
        ws['F8'] = total_student_takers
        # Total number of students count
        ws['G8'] = total_students
        # Overall Burnout Profiles count for all colleges
        get_overall_profile_counts(ws, overall_burnout_counts)
        # Burnout Profiles count per program for all colleges
        get_program_with_burnout_count(ws, program_info, profile_per_prog)
        # Burnout Profiles count per gender for all colleges
        get_profile_counts_per_gender(ws, profiles, profile_gender_counts)
        # Burnout Profiles count per age for all colleges
        get_profile_count_per_age(ws, selected_profiles, start_columns, start_row_age, remaining_profiles, start_rows_remaining, start_columns_remaining, age_data_per_profile)
        # Create new sheet for student assessment list for all colleges
        create_student_assessments_sheet(wb, student_assessments)
        # Create separate sheets for list of students according to profile for all colleges
        create_profile_sheets(wb, b_profiles, selected_college)



    # Create the in-memory Excel file
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Move the pointer to the beginning of the file
        
    # Create the HttpResponse object with the in-memory Excel file
    generated_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    response = HttpResponse(excel_file.read(), content_type="application/ms-excel")
    if selected_college:
        response['Content-Disposition'] = f'attachment; filename={selected_college}_Data_{generated_date}.xlsx'
    else:
        response['Content-Disposition'] = f'attachment; filename=Palawan_State_University_Data_{generated_date}.xlsx'
        
    # Close the in-memory Excel file
    excel_file.close()

    return response

# update the assessment per college
def enable_assessment(request):
    if request.method == 'POST':
        college_id = request.POST.get('college_id')
        is_enabled = request.POST.get('is_enabled')
        
        print("College ID: ", college_id)
        print("Is enabled? ", is_enabled)
        
        if is_enabled.lower() == 'true':
            is_enabled = True
        else:
            is_enabled = False
        
        try:
            college = College.objects.get(pk=college_id)
            college.is_assessment_enabled = is_enabled
            college.save()
            return JsonResponse({'message': 'College assessment status updated successfully.'})
        except College.DoesNotExist:
            return JsonResponse({'error': 'College not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    
 
# ----------------------------------------------------------------
#  ADMIN LIST OF STUDENTS PER BURNOUT PROFILES
# ----------------------------------------------------------------


@login_required(login_url="/login/")
def student_burnout_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(profile="Burned Out")
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        burnout_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        burnout_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        burnout_students = burnout_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "burnout_students": burnout_students,
    }

    return render(request, "admin/admin-burnedout-student.html", context)


@login_required(login_url="/login/")
def student_overextended_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(profile="Overextended")
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        overextended_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        overextended_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        overextended_students = overextended_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "overextended_students": overextended_students,
    }

    return render(request, "admin/admin-overextended-student.html", context)


@login_required(login_url="/login/")
def student_disengaged_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(profile="Disengaged")
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        disengaged_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        disengaged_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        disengaged_students = disengaged_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "disengaged_students": disengaged_students,
    }

    return render(request, "admin/admin-disengaged-student.html", context)


@login_required(login_url="/login/")
def student_ineffective_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(profile="Ineffective")
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        ineffective_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        ineffective_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        ineffective_students = ineffective_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "ineffective_students": ineffective_students,
    }

    return render(request, "admin/admin-ineffective-student.html", context)


@login_required(login_url="/login/")
def student_oandd_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(
        profile="Overextended and Disengaged"
    )
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        oandd_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        oandd_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        oandd_students = oandd_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "oandd_students": oandd_students,
    }

    return render(request, "admin/admin-oandd-student.html", context)


@login_required(login_url="/login/")
def student_oandi_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(
        profile="Overextended and Ineffective"
    )
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        oandi_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        oandi_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        oandi_students = oandi_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "oandi_students": oandi_students,
    }

    return render(request, "admin/admin-oandi-student.html", context)


@login_required(login_url="/login/")
def student_dandi_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(
        profile="Disengaged and Ineffective"
    )
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        dandi_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        dandi_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        dandi_students = dandi_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "dandi_students": dandi_students,
    }

    return render(request, "admin/admin-dandi-student.html", context)


@login_required(login_url="/login/")
def student_engaged_list(request, college_name=None):
    assessment_profile = BurnoutProfile.objects.filter(profile="Engaged")
    selected_college = None

    if college_name:
        selected_college = get_object_or_404(College, college_name=college_name)
        engaged_students = Assessment.objects.filter(
            student__program__college=selected_college, profile__in=assessment_profile
        ).order_by('-id')
    else:
        engaged_students = Assessment.objects.filter(profile__in=assessment_profile).order_by('-id')

    search = request.GET.get("q", "")

    if search:
        engaged_students = engaged_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {
        "search": search,
        "selected_college": selected_college,
        "engaged_students": engaged_students,
    }

    return render(request, "admin/admin-engaged-student.html", context)


# ----------------------------------------------------------------
#  COUNSERLOR SIDE FUNCTIONS
# ----------------------------------------------------------------

COUNSELOR_CACHE_PREFIX = 'counselor_dashboard_data'

# Counselor Dashboard
@login_required(login_url="/login/")
def counselor_dashboard(request):
    start_time = time.time()

    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    # Clean the college name
    counselor_college_cleaned = counselor_college.college_name.replace(' ', '_')  
    counselor_college_cleaned = ''.join(char for char in counselor_college_cleaned if char.isalnum() or char in ['_', '-'])  

    cache_key = f'{COUNSELOR_CACHE_PREFIX}_{counselor_college_cleaned}'

    # Check if the cache exists
    cached_data = cache.get(cache_key)

    new_assessment = check_for_new_assessment()

    if not cached_data or new_assessment:

        burnout_profiles_start_time = time.time()
        burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
        burnout_profiles_end_time = time.time()
        print(f"Burnout Profiles Query Time: {burnout_profiles_end_time - burnout_profiles_start_time} seconds")

        college_programs_start_time = time.time()
        college_programs = Program.objects.filter(college=counselor_college)
        college_programs_end_time = time.time()
        print(f"College Programs Query Time: {college_programs_end_time - college_programs_start_time} seconds")

        students_start_time = time.time()
        students = Student.objects.filter(program__college=counselor_college)
        students_end_time = time.time()
        print(f"Students Query Time: {students_end_time - students_start_time} seconds")
        
        student_assessments_start_time = time.time()
        student_assessments = Assessment.objects.select_related("student").filter(student__in=students).order_by('-pk')[:8]
        student_assessments_end_time = time.time()
        print(f"Student Assessments Query Time: {student_assessments_end_time - student_assessments_start_time} seconds")

        total_students_start_time = time.time()
        total_students = Student.objects.filter(program__college=counselor_college).count()
        total_students_end_time = time.time()
        print(f"Total Students Query Time: {total_students_end_time - total_students_start_time} seconds")
        
        total_student_takers_start_time = time.time()
        total_student_takers = Assessment.objects.select_related("student").filter(student__in=students).count()
        total_student_takers_end_time = time.time()
        print(f"Total Student Takers Query Time: {total_student_takers_end_time - total_student_takers_start_time} seconds")

        keywords_start_time = time.time()
        top_keywords = get_keywords_and_counts_for_students(students)
        keywords_end_time = time.time()
        print(f"Keywords Query Time: {keywords_end_time - keywords_start_time} seconds")

        overall_burnout_counts_start_time = time.time()
        overall_burnout_counts = get_overall_burnout_counts(counselor_college)
        overall_burnout_counts_end_time = time.time()
        print(f"Overall Burnout Counts Query Time: {overall_burnout_counts_end_time - overall_burnout_counts_start_time} seconds")
        
        gender_counts_start_time = time.time()
        profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)
        gender_counts_end_time = time.time()
        print(f"Gender Counts Query Time: {gender_counts_end_time - gender_counts_start_time} seconds")

        # Get program info specific to the logged-in counselor
        program_info_start_time = time.time()
        program_info = get_program_info(college_programs)
        program_info_end_time = time.time()
        print(f"Program Info Query Time: {program_info_end_time - program_info_start_time} seconds")

        chart_labels_start_time = time.time()
        chart_labels = list(profile_gender_counts.keys())
        chart_labels_end_time = time.time()
        print(f"Chart Labels Query Time: {chart_labels_end_time - chart_labels_start_time} seconds")

        male_data_start_time = time.time()
        male_data = [profile_gender_counts[profile]["Male"] for profile in chart_labels]
        male_data_end_time = time.time()
        print(f"Male Data Query Time: {male_data_end_time - male_data_start_time} seconds")
        
        female_data_start_time = time.time()
        female_data = [profile_gender_counts[profile]["Female"] for profile in chart_labels]
        female_data_end_time = time.time()
        print(f"Female Data Query Time: {female_data_end_time - female_data_start_time} seconds")

        context = {
            "counselor_college": counselor_college,
            "program_info": program_info,
            "overall_burnout_counts": overall_burnout_counts,
            "student_assessments": student_assessments,
            "total_students": total_students,
            "total_student_takers": total_student_takers,
            "chart_labels": chart_labels,
            "male_data": male_data,
            "female_data": female_data,
            "top_keywords": top_keywords,
            "age_data_per_profile": calculate_age_range_distribution(students, burnout_profiles),
        }

        cache.set(cache_key, context)

    else:
        context = cached_data

    end_time = time.time()
    print(f"Total Execution Time: {end_time - start_time} seconds")

    # return render(request, "counselor/counselor-dashboard.html", context)

    response = render(request, "counselor/counselor-dashboard.html", context)

    response['Cache-Control'] = 'max-age=' + str(6 * 30 * 24 * 60 * 60)  # Cache for 6 months
    response['Expires'] = (timezone.now() + timedelta(days=180)).strftime('%a, %d %b %Y %H:%M:%S GMT')

    return response


# Counselor all student list
@login_required(login_url="/login/")
def counselor_all_student_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    students = Student.objects.filter(program__college=counselor_college).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(email__icontains=search)
            | Q(program__program_name__icontains=search)
            | Q(assessment__profile__profile__icontains=search)
        ).distinct()


    context = {
        "students": students,
        "search": search,
    }

    return render(request, "counselor/counselor-all-student-list.html", context)


# Counselor student list
@login_required(login_url="/login/")
def counselor_student_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college


    students = Student.objects.filter(program__college=counselor_college, assessment_exists=False).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(email__icontains=search)
            | Q(program__program_name__icontains=search)
            | Q(assessment__profile__profile__icontains=search)
        ).distinct()


    context = {
        "students": students,
        "search": search,
    }

    return render(request, "counselor/counselor-student.html", context)


# Counselor student takers list
@login_required(login_url="/login/")
def counselor_student_takers_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    students = Student.objects.filter(program__college=counselor_college)

    search = request.GET.get("q", "")

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(email__icontains=search)
            | Q(program__program_name__icontains=search)
            | Q(assessment__profile__profile__icontains=search)
        ).distinct()

    student_assessments = (
        Assessment.objects.select_related("student")
        .filter(student__in=students)
        .order_by("-id")
    )

    context = {
        "student_assessments": student_assessments,
        "search": search,
    }

    return render(request, "counselor/counselor-student-takers.html", context)

# ----------------------------------------------------------------
#  COUNSELOR DOWNLOAD TO EXCEL FUNCTION
# ----------------------------------------------------------------

def counselor_download_excel(request):
    file_path = os.path.join(settings.STATIC_ROOT, 'report_temp', 'excel_dashboard.xlsx')

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college.college_name

    # Define the profiles variable with the profile names
    profiles = ['Engaged', 'Disengaged_and_Ineffective', 'Overextended_and_Ineffective', 'Overextended_and_Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned_out']
    b_profiles = ['Engaged', 'Disengaged and Ineffective', 'Overextended and Ineffective', 'Overextended and Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned Out']
    profile_per_prog = ['Engaged', 'Disengaged_and_Ineffective', 'Overextended_and_Ineffective', 'Overextended_and_Disengaged', 'Ineffective', 'Disengaged', 'Overextended', 'Burned_Out']

    # Age distribution data for selected profiles
    selected_profiles = ['Engaged', 'Disengaged and Ineffective', 'Overextended and Ineffective', 'Overextended and Disengaged']

    # Starting columns for each profile
    start_columns = {'Engaged': 'K', 'Disengaged and Ineffective': 'O', 'Overextended and Ineffective': 'S', 'Overextended and Disengaged': 'W'}
    start_row_age = 17

    # Age distribution data for remaining profiles (Ineffective, Disengaged, Overextended, Burned Out)
    remaining_profiles = ['Ineffective', 'Disengaged', 'Overextended', 'Burned Out']
    start_rows_remaining = 20

    # Starting columns for remaining profiles
    start_columns_remaining = {'Ineffective': 'K', 'Disengaged': 'O', 'Overextended': 'S', 'Burned Out': 'W'}

  
    college_programs = Program.objects.filter(college__college_name=counselor_college)
    students = Student.objects.filter(program__college__college_name=counselor_college)
    total_students = Student.objects.filter(program__college__college_name=counselor_college).count()
    total_student_takers = Assessment.objects.select_related("student").filter(
        student__program__college__college_name=counselor_college).count()
    student_assessments = Assessment.objects.select_related("student").filter(
        student__program__college__college_name=counselor_college).order_by('-created_date')
    burnout_profiles = BurnoutProfile.objects.all().exclude(pk=9)
    age_data_per_profile = calculate_age_range_distribution(students, burnout_profiles)
    program_info = get_program_info(college_programs)
    overall_burnout_counts = get_overall_burnout_counts(counselor_college)
    profile_gender_counts = get_profile_gender_counts(overall_burnout_counts)
    #-----------------------------------------------------------------
    # Populate to excel
    # ----------------------------------------------------------------
    
    # College title
    ws['B8'] = counselor_college 
    # Total student takers count
    ws['F8'] = total_student_takers
    # Total number of students count
    ws['G8'] = total_students
    # Overall Burnout Profiles count 
    get_overall_profile_counts(ws, overall_burnout_counts)
    # Burnout Profiles count per program
    get_program_with_burnout_count(ws, program_info, profile_per_prog)
    # Burnout Profiles count per gender
    get_profile_counts_per_gender(ws, profiles, profile_gender_counts)
    # Burnout Profiles count per age
    get_profile_count_per_age(ws, selected_profiles, start_columns, start_row_age, remaining_profiles, start_rows_remaining, start_columns_remaining, age_data_per_profile)
    # Create new sheet for student assessment list
    create_student_assessments_sheet(wb, student_assessments)
    # Create separate sheets for list of students according to profile
    create_profile_sheets(wb, b_profiles, counselor_college)

    # Create the in-memory Excel file
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Move the pointer to the beginning of the file
        
    # Create the HttpResponse object with the in-memory Excel file
    generated_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    response = HttpResponse(excel_file.read(), content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename={counselor_college}_Data_{generated_date}.xlsx'

    # Close the in-memory Excel file
    excel_file.close()

    return response
# ----------------------------------------------------------------
#  COUNSELOR LIST OF STUDENTS PER BURNOUT PROFILES
# ----------------------------------------------------------------

@login_required(login_url="/login/")
def counselor_student_burnout_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(profile="Burned Out")
    burnout_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        burnout_students = burnout_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "burnout_students": burnout_students}

    return render(request, "counselor/counselor-burnedout-student.html", context)


@login_required(login_url="/login/")
def counselor_student_overextended_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(profile="Overextended")
    overextended_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        overextended_students = overextended_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "overextended_students": overextended_students}

    return render(request, "counselor/counselor-overextended-student.html", context)


@login_required(login_url="/login/")
def counselor_student_disengaged_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(profile="Disengaged")
    disengaged_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        disengaged_students = disengaged_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "disengaged_students": disengaged_students}

    return render(request, "counselor/counselor-disengaged-student.html", context)


@login_required(login_url="/login/")
def counselor_student_ineffective_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(profile="Ineffective")
    ineffective_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        ineffective_students = ineffective_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "ineffective_students": ineffective_students}

    return render(request, "counselor/counselor-ineffective-student.html", context)


@login_required(login_url="/login/")
def counselor_student_oandd_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(
        profile="Overextended and Disengaged"
    )
    oandd_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        oandd_students = oandd_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "oandd_students": oandd_students}

    return render(request, "counselor/counselor-oandd-student.html", context)


@login_required(login_url="/login/")
def counselor_student_oandi_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(
        profile="Overextended and Ineffective"
    )
    oandi_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        oandi_students = oandi_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "oandi_students": oandi_students}

    return render(request, "counselor/counselor-oandi-student.html", context)


@login_required(login_url="/login/")
def counselor_student_dandi_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(
        profile="Disengaged and Ineffective"
    )
    dandi_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    search = request.GET.get("q", "")

    if search:
        dandi_students = dandi_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "dandi_students": dandi_students}

    return render(request, "counselor/counselor-dandi-student.html", context)


@login_required(login_url="/login/")
def counselor_student_engaged_list(request):
    counselor = Counselor.objects.get(account=request.user)
    counselor_college = counselor.college

    assessment_profile = BurnoutProfile.objects.filter(profile="Engaged")
    engaged_students = Assessment.objects.filter(
        student__program__college=counselor_college, profile__in=assessment_profile
    ).order_by("-id")

    # Get the search query from the request's GET parameters
    search = request.GET.get("q", "")

    # Filter students based on the search criteria
    if search:
        engaged_students = engaged_students.filter(
            Q(student__student_id__icontains=search)
            | Q(student__email__icontains=search)
            | Q(student__program__program_name__icontains=search)
        ).distinct()

    context = {"search": search, "engaged_students": engaged_students}

    return render(request, "counselor/counselor-engaged-student.html", context)


# ----------------------------------------------------------------
#  STUDENT SIDE FUNCTIONS
# ----------------------------------------------------------------


# Student Dashboard
@login_required(login_url="/login/")
def student_dashboard(request):
    student = Student.objects.filter(email=request.user.email).first()
    burnout_assessment = Assessment.objects.filter(student=student).first()

    # Check if the user has just logged in and the password is not updated
    # if request.session.get("just_logged_in", False) and not student.is_password_updated:
    # if not student.is_password_updated:
    #     messages.info(request, "Please update your profile information.")
        # # Remove the session variable to avoid showing the pop-up on subsequent visits
        # del request.session["just_logged_in"]

    # Specify parameters for YAKE 
    max_ngram_size = 5
    deduplication_threshold = 0.9
    numOfKeywords = 10

    custom_kw_extractor = yake.KeywordExtractor(
        lan="en",
        n=max_ngram_size,
        dedupLim=deduplication_threshold,
        top=numOfKeywords,
        features=None,
    )

    rake_extractor = Rake()

    # Define a list of indices for which RAKE will be used
    rake_indices = [3, 5, 6]

    print(f"RAKE will be used for instances with indices: {rake_indices}")

    # Define a mapping of statement indices to the desired keyword positions for RAKE
    rake_statement_mapping = {3: 0, 5: 0, 6: 0}

    # Define a mapping of statement indices to the desired keyword positions for YAKE
    yake_statement_mapping = {
        1: 9,
        2: 1,
        4: 0,
        7: 0,
        8: 0,
        9: 1,
        10: 0,
        11: 1,
        12: 1,
        13: 3,
        14: 0,
        15: 0,
    }  

    if burnout_assessment is not None:
        profile = burnout_assessment.profile.profile

        profile_to_questions = {
            "Burned Out": ["ex", "cy"],
            "Overextended": ["ex", "ef"],
            "Disengaged": ["cy", "ef"],
            "Ineffective": ["ex", "cy", "ef"],
            "Engaged": ["ef"],
            "Overextended and Disengaged": ["ex", "cy", "ef"],
            "Overextended and Ineffective": ["ex"],
            "Disengaged and Ineffective": ["cy"]
        }

        questions_to_extract = profile_to_questions.get(profile, [])

    # List to store the extracted keywords and scores
    extracted_keywords_and_scores = []

    # Retrieve the 15 most recent instances for the student
    recent_instances = StudentSurveyQuestion.objects.filter(student=student)[:15]

    for i, instance in enumerate(recent_instances, start=1):
        # Scrape text
        scraped_text = instance.scrape_question_text()
        question_code = instance.question.code.lower()

        if any(code in question_code for code in questions_to_extract):
            if i in rake_indices:  # Use RAKE for specific instances
                rake_extractor.extract_keywords_from_text(scraped_text)
                keywords = rake_extractor.get_ranked_phrases()

                # Choose a representative keyword based on the defined mapping for RAKE
                representative_keyword_position = rake_statement_mapping.get(i, 0)  # Default to the first keyword
            else:  # Use YAKE for other instances
                keywords_with_scores = custom_kw_extractor.extract_keywords(scraped_text)
                keywords = [keyword[0] for keyword in keywords_with_scores]  # Extract only the string part

                # Choose a representative keyword based on the defined mapping for YAKE
                representative_keyword_position = yake_statement_mapping.get(i, 0)  # Default to the first keyword

            # Filter out non-string keywords (including numpy.float64)
            keywords = [
                str(keyword)
                for keyword in keywords
                if isinstance(keyword, (str, np.str_))
                and not any(char.isdigit() for char in str(keyword))
            ]

            # Choose a representative keyword based on the defined mapping
            representative_keyword = keywords[representative_keyword_position] if keywords else "N/A"
            question_score = getattr(burnout_assessment, f"{question_code}_score", None)

            answer = str(instance.answer)
            result = re.sub(r'\([^)]*\)', '', answer)

            question_answer = result.strip()

            # Add the representative keyword and score to the list of extracted keywords and scores
            extracted_keywords_and_scores.append(
                {
                    "keywords": representative_keyword,
                    "score": question_score,
                    "answer": question_answer,
                }
            )

    # Sort the extracted keywords and scores by score in descending order
    sorted_extracted_keywords_and_scores = sorted(
        extracted_keywords_and_scores, key=lambda x: x["score"], reverse=True
    )

    # Select the top 5 highest-scoring questions
    top_5_keywords_and_scores = sorted_extracted_keywords_and_scores[:5]

    # Add the top 5 keywords and scores to the context
    context = {
        "student": student,
        "student_id": student.student_id,
        "burnout_assessment": burnout_assessment,
        "top_keywords": top_5_keywords_and_scores,
    }

    return render(request, "student/student-dashboard.html", context)


@login_required(login_url="/login/")
def student_edit_profile(request):
    student = Student.objects.filter(email=request.user.email).first()

    if request.method == "POST":
        form = UpdateProfileForm(request.POST, instance=student)
        if form.is_valid():
            form.save()

            # # Set the profile as updated 
            # if not student.is_profile_updated:
            #     student.is_profile_updated = True
            #     student.save()

            messages.success(request, "Your profile has been updated")
            return redirect("landingpage:edit-profile")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                error_messages.append(f"Error in {field}: {', '.join(errors)}")
            error_message = "Failed to update your profile. " + " ".join(error_messages)
            messages.error(request, error_message)

    else:
        form = UpdateProfileForm(instance=student)

    return render(
        request,
        "student/edit-profile.html",
        {"form": form, "student": student, "student_id": student.student_id},
    )


# ----------------------------------------------------------------
#  ADMIN AND COUNSELOR SHARED FUNCTIONS
# ----------------------------------------------------------------


# Admin and Counselor Student Profile
@login_required(login_url="/login/")
def student_profile(request, student_id):
    student = get_object_or_404(Student, student_id=student_id)
    assessment = Assessment.objects.filter(student=student).first()

    if assessment:
        burnout_profile = BurnoutProfile.objects.get(id=assessment.profile.id)

        # Fetch the student's survey answers (limit to the most recent 15 questions based on PK)
        student_survey_questions = StudentSurveyQuestion.objects.filter(
            student=student
        ).order_by("-pk")[:15]

        # Create a mapping between the question codes and score attribute names
        code_to_score = {
            "EX1": "ex1_score",
            "EX2": "ex2_score",
            "EX3": "ex3_score",
            "EX4": "ex4_score",
            "EX5": "ex5_score",
            "CY1": "cy1_score",
            "CY2": "cy2_score",
            "CY3": "cy3_score",
            "CY4": "cy4_score",
            "EF1": "ef1_score",
            "EF2": "ef2_score",
            "EF3": "ef3_score",
            "EF4": "ef4_score",
            "EF5": "ef5_score",
            "EF6": "ef6_score",
        }

        # Extract the scores 
        scores = []

        for survey_question in student_survey_questions:
            # Get the code of the question
            question_code = survey_question.question.code

            # Map the question code to the corresponding score attribute name
            score_attribute = code_to_score.get(question_code, None)

            if score_attribute is not None:
                score_value = getattr(assessment, score_attribute, None)
                scores.append(score_value)

        # Serialize score data to JSON format
        scores_json = json.dumps(scores)

        # Extract the student's survey answers and question texts
        survey_answers = [
            survey_question.answer.choice
            for survey_question in student_survey_questions
        ]
        question_texts = [
            survey_question.question.question
            for survey_question in student_survey_questions
        ]

        answers_json = json.dumps(survey_answers)
        questions_json = json.dumps(question_texts)

        context = {
            "student": student,
            "assessment": assessment,
            "burnout_profile": burnout_profile,
            "survey_answers": answers_json,
            "survey_questions": questions_json,
            "student_survey": student_survey_questions,
            "scores": scores_json,
        }

        return render(request, "admin-counselor/student-profile.html", context)


# ----------------------------------------------------------------
#  TEMPLATE VIEW
# ----------------------------------------------------------------


# IBPS TERMS OF SERVICE
def tos(request):
    return render(request, "tos.html")


# IBPS PRIVACY POLICY
def privacy_policy(request):
    return render(request, "privacy_policy.html")


# STUDENT ASSESSMENT PAGE 1
@login_required(login_url="/login/")
def student_assessment1(request):
    student = Student.objects.filter(email=request.user.email).first()
    college = College.objects.filter(college_name=student.program.college).first()

    return render( request, "student/assessment-1.html", {"student": student, "student_id": student.student_id, "college": college})


# STUDENT ASSESSMENT PAGE 2
@login_required(login_url="/login/")
def student_assessment2(request):
    student = Student.objects.filter(email=request.user.email).first()
    return render(
        request,
        "student/assessment-2.html",
        {"student": student, "student_id": student.student_id},
    )


# ----------------------------------------------------------------
# OTHER FUNCTIONS
# ----------------------------------------------------------------


def get_counselor_email(request, student_id):
    try:
        student = Student.objects.get(id=student_id)
        counselor = Counselor.objects.get(college=student.program.college)
        return JsonResponse({"counselor_email": counselor.email})
    except (Student.DoesNotExist, Counselor.DoesNotExist) as e:
        return JsonResponse({"error": str(e)}, status=400)
    

def get_program_id_mapping():
    programs = Program.objects.all()
    
    program_id_mapping = {program.program_name: program.id for program in programs}
    
    return program_id_mapping

# ----------------------------------------------------------------
# DOWNLOAD TO EXCEL FUNCTIONS
# ----------------------------------------------------------------

def add_image_with_text(ws_students, text1, text2):
    image_path = os.path.join(settings.STATIC_ROOT, 'img', 'psu', 'PSU LOGO.png')

    img = Image(image_path)
    ws_students.add_image(img, 'D2')
    ws_students['E3'] = text1
    ws_students['E4'] = text2
    bold_font = Font(bold=True)
    ws_students['E3'].font = bold_font
    ws_students['E4'].font = bold_font
    alignment = Alignment(horizontal='center')
    ws_students['E3'].alignment = alignment
    ws_students['E4'].alignment = alignment


def get_overall_profile_counts(ws, overall_burnout_counts):
    # Iterate over the 8 columns from B to I
    for idx, (profile_name, counts_dict) in enumerate(overall_burnout_counts.items()):
        column = chr(ord('B') + idx)  # Convert index to corresponding Excel column (B, C, D, ...)
        ws[f'{column}11'] = counts_dict.get('overall_count', 0)


def get_program_with_burnout_count(ws, program_info, profile_per_prog):
    start_column = 'B'
    end_column = 'I'
    num_columns = ord(end_column) - ord(start_column) + 1

    for i, program_data in enumerate(program_info):
        program_name = program_data['program'].program_name
        ws[f'A{15+i}'] = program_name

        # Insert burnout counts for each program
        for j in range(num_columns):
            profile_count = program_data['burnout_counts'].get(profile_per_prog[j], 0)
            column = chr(ord(start_column) + j) 
            ws[f'{column}{15+i}'] = profile_count


def get_profile_counts_per_gender(ws, profiles, profile_gender_counts):
    # Add gender counts for each profile
    for idx, profile in enumerate(profiles):
        male_count = profile_gender_counts.get(profile, {}).get("Male", 0)
        female_count = profile_gender_counts.get(profile, {}).get("Female", 0)
            
        # Set the column letters for male and female data
        male_column = chr(ord('K') + idx * 2)
        female_column = chr(ord('L') + idx * 2)
            
        # Set the values for male and female counts
        ws[f'{male_column}{10}'] = male_count  # Male count for the profile
        ws[f'{female_column}{10}'] = female_count  # Female count for the profile


def get_profile_count_per_age(ws, selected_profiles, start_columns, start_row_age, remaining_profiles, start_rows_remaining, start_columns_remaining, age_data_per_profile):
    # Iterate over the selected profiles
    for profile_name in selected_profiles:
        start_column = start_columns.get(profile_name)
        if start_column:
            # Write profile name
            ws[f'{start_column}{start_row_age}'] = profile_name
                
            # Write age distribution data for each age range
            age_data = age_data_per_profile.get(profile_name, [])
            for age_idx, count in enumerate(age_data):
                column = chr(ord(start_column) + age_idx)  # Get the column letter
                ws[f'{column}{start_row_age}'] = count


    # Iterate over the remaining profiles
    for profile_name in remaining_profiles:
        start_column = start_columns_remaining.get(profile_name)
        if start_column:
            # Write profile name
            ws[f'{start_column}{start_rows_remaining}'] = profile_name
                
            # Write age distribution data for each age range
            age_data = age_data_per_profile.get(profile_name, [])
            for age_idx, count in enumerate(age_data):
                column = chr(ord(start_column) + age_idx)  # Get the column letter
                ws[f'{column}{start_rows_remaining}'] = count


def create_student_assessments_sheet(wb, student_assessments):
    # Create a new sheet named "Student Assessments"
    ws_students = wb.create_sheet("Student Assessments")

    add_image_with_text(ws_students, 'Republic of the Philippines', 'Palawan State University')

    # Headers for the student assessments
    headers = ['Date of Examination', 'Student Name', 'Student ID', 'Email', 'Contact Number', 'College', 'Program', 'Burnout Profile']
    for idx, header in enumerate(headers, start=1):
        cell = ws_students.cell(row=7, column=idx, value=header)
        cell.font = Font(bold=True)  
        cell.alignment = Alignment(horizontal='center')

    # Populate the sheet with student assessment data
    for i, student_assessment in enumerate(student_assessments):
        last_name = student_assessment.student.last_name
        first_name = student_assessment.student.first_name
        middle_name = student_assessment.student.middle_name or ""

        full_name = ''
        if last_name and first_name:
            full_name = f"{last_name.upper()}, {first_name}"
        if middle_name:
            full_name += f" {middle_name}"

        date_of_examination = student_assessment.created_date.astimezone(pytz.utc).replace(tzinfo=None)
        student_name = full_name
        student_id = student_assessment.student.student_id
        student_email = student_assessment.student.email
        student_cnum = student_assessment.student.contact_number
        student_college = student_assessment.student.program.college.college_name
        student_program = student_assessment.student.program.program_name
        student_bprofile = student_assessment.profile.profile
            
        # Write data to the "Student Assessments" sheet
        ws_students.cell(row=8+i, column=1, value=date_of_examination)
        ws_students.cell(row=8+i, column=2, value=student_name)
        ws_students.cell(row=8+i, column=3, value=student_id)
        ws_students.cell(row=8+i, column=4, value=student_email)
        ws_students.cell(row=8+i, column=5, value=student_cnum)
        ws_students.cell(row=8+i, column=6, value=student_college)
        ws_students.cell(row=8+i, column=7, value=student_program)
        ws_students.cell(row=8+i, column=8, value=student_bprofile)

    return ws_students


def create_profile_sheets(wb, b_profiles,  selected_college):
    for profile in b_profiles:
        # Create a new sheet with the profile name
        ws_profile = wb.create_sheet(profile)

        add_image_with_text(ws_profile, 'Republic of the Philippines', 'Palawan State University')

        # Headers for the profile sheet
        headers = ['Date of Examination', 'Student Name', 'Student ID', 'Email', 'Contact Number', 'College', 'Program']
        for idx, header in enumerate(headers, start=1):
            cell = ws_profile.cell(row=7, column=idx, value=header)
            cell.font = Font(bold=True)  
            cell.alignment = Alignment(horizontal='center')

        # Get students with the current profile
        if selected_college:
            students_with_profile = Assessment.objects.filter(profile__profile=profile, student__program__college__college_name=selected_college).order_by("-created_date")
        else:
            students_with_profile = Assessment.objects.filter(profile__profile=profile).order_by("-created_date")

        # Populate the sheet with student data
        for i, assessment in enumerate(students_with_profile):
            last_name = assessment.student.last_name
            first_name = assessment.student.first_name
            middle_name = assessment.student.middle_name or ""

            full_name = ''
            if last_name and first_name:
                full_name = f"{last_name.upper()}, {first_name}"
            if middle_name:
                full_name += f" {middle_name}"

            date_of_examination = assessment.created_date.astimezone(pytz.utc).replace(tzinfo=None)
            student_name = full_name
            student_id = assessment.student.student_id
            student_email = assessment.student.email
            student_cnum = assessment.student.contact_number
            student_college = assessment.student.program.college.college_name
            student_program = assessment.student.program.program_name


            ws_profile.cell(row=8+i, column=1, value=date_of_examination)
            ws_profile.cell(row=8+i, column=2, value=student_name)
            ws_profile.cell(row=8+i, column=3, value=student_id)
            ws_profile.cell(row=8+i, column=4, value=student_email)
            ws_profile.cell(row=8+i, column=5, value=student_cnum)
            ws_profile.cell(row=8+i, column=6, value=student_college)
            ws_profile.cell(row=8+i, column=7, value=student_program)

    return wb


# Calculate student age based on birthdate
def calculate_age(student_id):
    student = Student.objects.get(student_id=student_id)

    current_date = datetime.date.today()

    if (current_date.month, current_date.day) >= (student.birthdate.month, student.birthdate.day):
        age = current_date.year - student.birthdate.year
    else:
        age = current_date.year - student.birthdate.year - 1

    student.age = age
    student.save()



# # Counselor burnout factors chart
# def get_keywords_and_counts(request):
#     counselor_college = request.user.counselor.college
#     students = Student.objects.filter(program__college=counselor_college)

#     max_ngram_size_yake = 5
#     deduplication_threshold_yake = 0.9
#     numOfKeywords_yake = 10

#     custom_kw_extractor_yake = yake.KeywordExtractor(lan="en", n=max_ngram_size_yake, dedupLim=deduplication_threshold_yake, top=numOfKeywords_yake, features=None)
#     rake_extractor = Rake()

#     rake_indices = [3, 5, 6]
#     rake_statement_mapping = {3: 0, 5: 0, 6: 0}
#     yake_statement_mapping = {1: 9, 2: 1, 4: 0, 7: 0, 8: 0, 9: 3, 10: 0, 11: 3, 12: 1, 13: 5, 14: 0, 15: 0}

#     extracted_keywords_and_scores = []

#     for student in students:
#         recent_instances = StudentSurveyQuestion.objects.filter(student=student)[:15]
#         burnout_assessment = Assessment.objects.filter(student=student).first()

#         for i, instance in enumerate(recent_instances, start=1):
#             scraped_text = instance.scrape_question_text()

#             if i in rake_indices:
#                 rake_extractor.extract_keywords_from_text(scraped_text)
#                 keywords = rake_extractor.get_ranked_phrases()
#                 representative_keyword_position = rake_statement_mapping.get(i, 0)
#             else:
#                 keywords_with_scores = custom_kw_extractor_yake.extract_keywords(scraped_text)
#                 keywords = [keyword[0] for keyword in keywords_with_scores]
#                 representative_keyword_position = yake_statement_mapping.get(i, 0)

#             keywords = [str(keyword) for keyword in keywords if isinstance(keyword, (str, np.str_)) and not any(char.isdigit() for char in str(keyword))]
#             representative_keyword = keywords[representative_keyword_position] if keywords and 0 <= representative_keyword_position < len(keywords) else "N/A"

#             question_code = instance.question.code.lower()
#             question_score = getattr(burnout_assessment, f'{question_code}_score', None)

#             # Skip instances where question_score is None
#             if question_score is not None:
#                 extracted_keywords_and_scores.append({'keywords': representative_keyword, 'score': question_score})


#     # Filter keywords with the highest scores (5 and 6)
#     highest_score_keywords = [entry['keywords'] for entry in extracted_keywords_and_scores if entry['score'] in {5, 6}]

#     #     # Extracted keywords
#     # extracted_keywords = [entry['keywords'] for entry in extracted_keywords_and_scores]

#     # Count occurrences of each unique keyword
#     keyword_counts = dict(Counter(highest_score_keywords))
#     print('Keyword count:', keyword_counts)

#     # Get the top 15 unique keywords
#     top_keywords = dict(sorted(keyword_counts.items(), key=lambda item: item[1], reverse=True)[:15])

#     print('Extracted keywords:', extracted_keywords_and_scores)
#     print('Top Keywords:', top_keywords)


#     return JsonResponse(top_keywords, safe=False)

